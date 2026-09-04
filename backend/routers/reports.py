import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
from database import db
from auth import require_permission

router = APIRouter()

# Report definitions: key -> (title, columns [(header, field)], collection loader)
ASSET_COLS = [("Nama Barang", "nama_barang"), ("Kode Barang", "kode_barang"), ("NUP", "nup"),
              ("Merk/Tipe", "merk_tipe"), ("Tahun", "tahun_perolehan"), ("Kondisi", "kondisi"),
              ("Lokasi", "lokasi"), ("Penanggung Jawab", "penanggung_jawab"), ("Status", "status")]
VEHICLE_COLS = [("Nama", "nama_barang"), ("No. Polisi", "nomor_polisi"), ("Jenis", "jenis_kendaraan"),
                ("No. Rangka", "nomor_rangka"), ("No. Mesin", "nomor_mesin"), ("Kondisi", "kondisi"),
                ("Penanggung Jawab", "penanggung_jawab")]
MAINT_COLS = [("No. Pengajuan", "request_number"), ("Aset", "asset_name"), ("No. Polisi", "nomor_polisi"),
              ("Jenis Pemeliharaan", "jenis_pemeliharaan"), ("Pemohon", "created_by_name"),
              ("Status", "status"), ("Perkiraan Biaya", "perkiraan_biaya")]


async def _load(report: str, filters: dict):
    q = {}
    if report in ("aset", "kondisi", "lokasi", "penanggung_jawab"):
        if filters.get("kondisi"):
            q["kondisi"] = filters["kondisi"]
        if filters.get("lokasi"):
            q["lokasi"] = filters["lokasi"]
        data = await db.assets.find(q, {"_id": 0}).to_list(5000)
        return "Laporan Data Aset BMN", ASSET_COLS, data
    if report == "kendaraan":
        q["jenis_aset"] = "Kendaraan"
        if filters.get("kondisi"):
            q["kondisi"] = filters["kondisi"]
        data = await db.assets.find(q, {"_id": 0}).to_list(5000)
        return "Laporan Kendaraan Dinas", VEHICLE_COLS, data
    if report == "pemeliharaan":
        if filters.get("status"):
            q["status"] = filters["status"]
        data = await db.maintenance_requests.find(q, {"_id": 0}).to_list(5000)
        return "Laporan Pemeliharaan", MAINT_COLS, data
    if report in ("barang", "barang_subsi"):
        if filters.get("status"):
            q["status"] = filters["status"]
        if filters.get("unit"):
            q["unit"] = filters["unit"]
        data = await db.inventory_requests.find(q, {"_id": 0}).to_list(5000)
        # flatten items
        rows = []
        for r in data:
            for it in r.get("items", []):
                rows.append({"request_number": r.get("request_number"), "pemohon_name": r.get("pemohon_name"),
                             "unit": r.get("unit"), "nama_barang": it.get("nama_barang"),
                             "jumlah": it.get("jumlah"), "satuan": it.get("satuan"),
                             "keperluan": it.get("keperluan"), "status": r.get("status")})
        cols = [("No. Permintaan", "request_number"), ("Pemohon", "pemohon_name"), ("Unit", "unit"),
                ("Nama Barang", "nama_barang"), ("Jumlah", "jumlah"), ("Satuan", "satuan"),
                ("Keperluan", "keperluan"), ("Status", "status")]
        title = "Laporan Barang per Subsi/Unit" if report == "barang_subsi" else "Laporan Permintaan Barang Persediaan"
        return title, cols, rows
    raise HTTPException(status_code=400, detail="Jenis laporan tidak dikenal")


@router.get("/reports/data")
async def report_data(report: str, kondisi: Optional[str] = None, lokasi: Optional[str] = None,
                      status: Optional[str] = None, unit: Optional[str] = None,
                      user=Depends(require_permission("reports_view"))):
    filters = {"kondisi": kondisi, "lokasi": lokasi, "status": status, "unit": unit}
    title, cols, data = await _load(report, filters)
    return {"title": title, "columns": [{"header": h, "field": f} for h, f in cols], "rows": data}


@router.get("/reports/export/excel")
async def export_excel(report: str, kondisi: Optional[str] = None, lokasi: Optional[str] = None,
                       status: Optional[str] = None, unit: Optional[str] = None,
                       user=Depends(require_permission("reports_view"))):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    filters = {"kondisi": kondisi, "lokasi": lokasi, "status": status, "unit": unit}
    title, cols, data = await _load(report, filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for ci, (h, f) in enumerate(cols, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
    for ri, row in enumerate(data, 4):
        for ci, (h, f) in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=row.get(f, ""))
    for ci in range(1, len(cols) + 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26 else "A"].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{report}_laporan.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/reports/export/pdf")
async def export_pdf(report: str, kondisi: Optional[str] = None, lokasi: Optional[str] = None,
                     status: Optional[str] = None, unit: Optional[str] = None,
                     user=Depends(require_permission("reports_view"))):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    filters = {"kondisi": kondisi, "lokasi": lokasi, "status": status, "unit": unit}
    title, cols, data = await _load(report, filters)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                            leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f"<b>{title}</b>", styles["Title"]),
                Paragraph("Sistem Monitoring BMN dan Barang Persediaan", styles["Normal"]),
                Spacer(1, 12)]
    table_data = [[h for h, f in cols]]
    for row in data:
        table_data.append([str(row.get(f, "") or "") for h, f in cols])
    if len(table_data) == 1:
        table_data.append(["Tidak ada data"] + [""] * (len(cols) - 1))
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={report}_laporan.pdf"})
